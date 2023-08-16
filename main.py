import cv2
import torch
import numpy as np
from tracker import *
from tqdm import tqdm
import openpyxl

# Khai bao mau
color_xe_may = (254,226,102)    #xanh ngoc
color_xe_bus = (102,236,254)    #vang
color_xe_oto = (0,243,104)      #la cay
color_xe_tai = (9, 9, 237)       #do
color_roi    = (184, 19, 193)   #tim

# Khai bao toa do
toado_1 = (850, 280)
toado_2 = (200, 280)
toado_3 = (0, 600)
toado_4 = (1080, 600)
toado_check_1 = (50, 520)
toado_check_2 =  (1020, 520)
area1 = [toado_1, toado_2, toado_3, toado_4]

#khai bao bien
count = 0  
xe_bus = []
xe_oto = []
xe_may = []
xe_tai = []
tracker = Tracker()



# Hàm vẽ khung
def draw_rectangle(frame, x1, y1, x2, y2, color, name):
    cv2.rectangle(frame, (x1, y1), (x2, y2), color, thickness=2)
    cv2.putText(frame, f"{name} ID:{id} ", (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, thickness=1)
    cv2.circle(frame, point, 1, color, thickness=2 )

# hàm check và đếm
def check_count(frame, toado, check_1, check_2, id, vehicle_list, conf ):
    if (toado[1] >= check_1[1] and toado[0] >= check_1[0] and toado[0] <= check_2[0]):
        if id not in vehicle_list:
            vehicle_list.append(id)

# Xuất dữ liệu đi qua roi     
def output_roi(frame, toado, x3, y3, x4, y4, check_1, check_2, id, name, conf, name_test):
    if toado[1] >= check_1[1] and toado[0] >= check_1[0] and toado[0] <= check_2[0]:
        try:
            wb = openpyxl.load_workbook('output/output_roi_{}.xlsx'.format(name_test))
            sheet = wb.active
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.append(["Toạ độ x1", "Toạ độ y1", "Toạ độ x2", "Toạ độ y2", "Loại Phương Tiện ", "ID", "Độ chính xác"])

        data_row = [x3, y3, x4, y4, name, id, conf]
        sheet.append(data_row)
        wb.save('output/output_roi_{}.xlsx'.format(name_test))

def output_total(xe_may, xe_oto,xe_bus, xe_tai, tong_xe, name_test):
    wb_total = openpyxl.Workbook()
    sheet_total = wb_total.active
    sheet_total.append(["Xe_may", "Xe_oto", "Xe_bus", "Xe_tai", "Tong_xe"])
    data_row_total = [xe_may, xe_oto,xe_bus, xe_tai, tong_xe]
    sheet_total.append(data_row_total)
    wb_total.save('output/output_total_{}.xlsx'.format(name_test))
     

#load model nhận diện
model = torch.hub.load('ultralytics/yolov5', 'custom', path='best.pt', force_reload=True) 


name_test = "test"
video_path_input = 'input/{}.mp4'.format(name_test)
video_path_output = 'output/{}.mp4'.format(name_test)

#đầu vào video
cap=cv2.VideoCapture(video_path_input)

#đầu ra video
video_codec = cv2.VideoWriter_fourcc(*"mp4v")  # Codec video (sử dụng MP4V)
output_video = cv2.VideoWriter(video_path_output, video_codec, 20, (1280, 800))

n_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
progress_bar = tqdm(total=n_frames)

# def POINTS(event, x, y, flags, param):
#     if event == cv2.EVENT_MOUSEMOVE :  
#         colorsBGR = [x, y]
#         print(colorsBGR)
        

# cv2.namedWindow('FRAME')
# cv2.setMouseCallback('FRAME', POINTS)

while True:
    ret,frame=cap.read()
    if not ret:
        break
    count += 1
    if count % 3 != 0:
        continue

    
    frame=cv2.resize(frame,(1280,800))
    results=model(frame)
    
    list = []

    #vẽ khu vực nhận diện
    cv2.polylines(frame, [np.array(area1, np.int32)], True, color_roi , 2)
    #đường roi để đếm đối tượng
    cv2.line(frame, toado_check_1, toado_check_2, (0,255,192), thickness=2)

    #lấy các giá trị sau khi nhận diện 
    for index, rows in results.pandas().xyxy[0].iterrows():
        x1 = int(rows['xmin'])
        y1 = int(rows['ymin'])
        x2 = int(rows['xmax'])
        y2 = int(rows['ymax'])
        b = str(rows['name'])
        conf= str(rows['confidence'])
        clas = rows['class']
        list.append([x1, y1, x2, y2, clas, conf])
    idx_bbox = tracker.update(list)

    # lấy lại giá trị sau khi thực hiện track đối tượng và xử lý
    for bbox in idx_bbox:
        x3, y3, x4, y4, id, clas, conf = bbox
        x_midpoint = int((x3+x4)/2)
        y_midpoint = int((y3+y4)/2)
        point = (x_midpoint, y_midpoint)
        if cv2.pointPolygonTest(np.array(area1, np.int32), point, False) >= 0:
    

            if clas == 1: #xe_bus
                name = "xe_bus"
                color = color_xe_bus
                draw_rectangle(frame, x3, y3, x4, y4, color, name)
                check_count(frame, point, toado_check_1, toado_check_2, id, xe_bus, conf)
                output_roi(frame, point, x3, y3, x4, y4, toado_check_1, toado_check_2, id, name, conf,name_test)
                

            if clas == 2: #oto
                name = "xe_oto"
                color = color_xe_oto
                draw_rectangle(frame, x3, y3, x4, y4, color, name)
                check_count(frame, point, toado_check_1, toado_check_2, id, xe_oto, conf)
                output_roi(frame, point, x3, y3, x4, y4, toado_check_1, toado_check_2, id, name, conf,name_test)
                    
            if clas == 3: #xe_may
                name = "xe_may"
                color = color_xe_may
                draw_rectangle(frame, x3, y3, x4, y4, color, name)
                check_count(frame, point, toado_check_1, toado_check_2, id, xe_may, conf)
                output_roi(frame, point, x3, y3, x4, y4, toado_check_1, toado_check_2, id, name, conf,name_test)
                

            if clas == 4: #xe_tai
                name = "xe_tai"
                color = color_xe_tai
                draw_rectangle(frame, x3, y3, x4, y4, color, name)
                check_count(frame, point, toado_check_1, toado_check_2, id, xe_tai, conf)
                output_roi(frame, point, x3, y3, x4, y4, toado_check_1, toado_check_2, id, name, conf,name_test)
                
        else:
            pass
        
        cv2.putText(frame, f"Xe_oto :{len(xe_oto)} ", (0, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, color_xe_oto, thickness=2)
        cv2.putText(frame, f"Xe_may :{len(xe_may)} ", (0, 70), cv2.FONT_HERSHEY_SIMPLEX, 1, color_xe_may, thickness=2)
        cv2.putText(frame, f"Xe_bus :{len(xe_bus)} ", (0, 110), cv2.FONT_HERSHEY_SIMPLEX, 1, color_xe_bus, thickness=2)
        cv2.putText(frame, f"Xe_tai :{len(xe_tai)} ", (0, 150), cv2.FONT_HERSHEY_SIMPLEX, 1, color_xe_tai, thickness=2)
        total = len(xe_oto) + len(xe_may) +len(xe_bus) + len(xe_tai)
        cv2.putText(frame, f"Tong_xe :{total} ", (0, 190), cv2.FONT_HERSHEY_SIMPLEX, 1, color_roi, thickness=2)

        output_total(len(xe_may), len(xe_oto), len(xe_bus), len(xe_tai), total,name_test)

    progress_bar.update(1)
    output_video.write(frame)

    # cv2.imshow("FRAME",frame)
    # if cv2.waitKey(1)&0xFF==27:
    #     break

    


cap.release()
progress_bar.close()
output_video.release()
cv2.destroyAllWindows()
